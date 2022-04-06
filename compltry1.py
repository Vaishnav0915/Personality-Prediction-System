# -*- coding: utf-8 -*-
"""
@author: vadre
"""

import pandas as pd
from numpy import *
import numpy as np
from sklearn import preprocessing
from sklearn import datasets, linear_model
from sklearn.metrics import mean_squared_error, r2_score
from sklearn import metrics
#from sklearn.cross_validation import train_test_split
from sklearn.model_selection import train_test_split
from sklearn import neighbors
from openpyxl import load_workbook
import sys
import os


# p1=r"C:\Users\vadre\Documents\adt project\new.xlsx"
# p2=r"C:\Users\vadre\Documents\adt project\train.csv"
t1=pd.read_excel(r"C:\Users\vadre\Documents\adt project\new.xlsx")
print(t1)
#print(t1.iloc[:,16])
#print(t1.columns[16])
t1.insert(14,'openness','')
t1.insert(25,'neuroticism','')
t1.insert(36,'conscientiousness','')
t1.insert(47,'agreeableness','')
t1.insert(58,'extraversion','')
t1.insert(60,'Personality (class label)','')
#print(t1.columns[62])
#adding the values of individual questions in order to get a characterisitic value
t1.iloc[:,14]=(t1.iloc[:,4:14].sum(axis=1))/10
t1.iloc[:,25]=(t1.iloc[:,15:25].sum(axis=1))/10
t1.iloc[:,36]=(t1.iloc[:,26:36].sum(axis=1))/10
t1.iloc[:,47]=(t1.iloc[:,37:47].sum(axis=1))/10
t1.iloc[:,58]=(t1.iloc[:,48:58].sum(axis=1))/10
#chanignt the data type to int from float
t1.iloc[:,14]=t1.iloc[:,14].astype(int)
t1.iloc[:,25]=t1.iloc[:,25].astype(int)
t1.iloc[:,36]=t1.iloc[:,36].astype(int)
t1.iloc[:,47]=t1.iloc[:,47].astype(int)
t1.iloc[:,58]=t1.iloc[:,58].astype(int)
    
t1.loc[t1.iloc[:,59] == 1, 'Personality (class label)'] = 'responsible'
t1.loc[t1.iloc[:,59] == 2, 'Personality (class label)'] = 'extraverted'
t1.loc[t1.iloc[:,59] == 3, 'Personality (class label)'] = 'serious'
t1.loc[t1.iloc[:,59] == 4, 'Personality (class label)'] = 'lively'
t1.loc[t1.iloc[:,59] == 5, 'Personality (class label)'] = 'dependable'
#chars=['lively','dependable','serious','responsible','extraverted']
#t1['Personality (class label)'] = np.random.shuffle(chars)
print(t1)
t1.to_excel("o1.xlsx")
td=load_workbook("o1.xlsx")
tds=td["Sheet1"]
tds.delete_cols(1,3)
tds.delete_cols(3,10)
tds.delete_cols(4,10)
tds.delete_cols(5,10)
tds.delete_cols(6,10)
tds.delete_cols(7,10)
tds.delete_cols(8)
tds.delete_cols(9,2)
td.save('td.xlsx')
testdata=pd.read_excel('td.xlsx')
testdata.to_csv('test dataset.csv',index=False)
os.remove('o1.xlsx')
os.remove('td.xlsx')
    
    
data =pd.read_csv(r"C:\Users\vadre\Documents\adt project\train dataset.csv")
array = data.values
for i in range(len(array)):
    if array[i][0]=="Male":
        array[i][0]=1
    else:
        array[i][0]=0
            
            
df=pd.DataFrame(array)
maindf =df[[0,1,2,3,4,5,6]]
mainarray=maindf.values
print (mainarray)
temp=df[7]
train_y =temp.values
# print(train_
# print(mainarray)
  
train_y=temp.values
for i in range(len(train_y)):
    train_y[i] =str(train_y[i])
mul_lr = linear_model.LogisticRegression(multi_class='multinomial', solver='newton-cg',max_iter =1000)
mul_lr.fit(mainarray, train_y)
testdata =pd.read_csv('test dataset.csv')
test = testdata.values
for i in range(len(test)):
    if test[i][0]=="Male":
        test[i][0]=1
    else:
        test[i][0]=0
df1=pd.DataFrame(test)
testdf =df1[[0,1,2,3,4,5,6]]
maintestarray=testdf.values
print(maintestarray)
y_pred = mul_lr.predict(maintestarray)
for i in range(len(y_pred)) :
    y_pred[i]=str((y_pred[i]))
DF = pd.DataFrame(y_pred,columns=['Predicted Personality'])
DF.index=DF.index+1
DF.index.names = ['Person No']
DF.to_csv(r"C:\Users\vadre\Documents\adt project\opt.csv")




eh=pd.read_csv("opt.csv")
n1=pd.read_csv(r"C:\Users\vadre\Documents\adt project\opt.csv")
n1.to_excel(r"C:\Users\vadre\Documents\adt project\eh1.xlsx", index=False)
df2=pd.DataFrame()
eh1=load_workbook(r"C:\Users\vadre\Documents\adt project\new.xlsx")
ehs1=eh1["Sheet1"]
ehs1.delete_cols(1,1)
ehs1.delete_cols(4,53)
eh1.save(r"C:\Users\vadre\Documents\adt project\new1.xlsx")
eh2=load_workbook(r"C:\Users\vadre\Documents\adt project\eh1.xlsx")
ehs2=eh2["Sheet1"]
ehs2.delete_cols(1,1)
eh2.save(r"C:\Users\vadre\Documents\adt project\eh1-1.xlsx")
n1=pd.read_excel(r"C:\Users\vadre\Documents\adt project\new1.xlsx")
n2=pd.read_excel(r"C:\Users\vadre\Documents\adt project\eh1-1.xlsx")
df=pd.concat([n1,n2],axis=1)
df.to_csv(r"C:\Users\vadre\Documents\adt project\opt2.csv",index=False)



 
    